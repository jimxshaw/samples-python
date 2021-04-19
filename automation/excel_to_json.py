import openpyxl
import os
import json

os.chdir('/Users/jimxshaw/Projects/samples-python/automation')


workbook = openpyxl.load_workbook('splunk_v4.19.0_cim.xlsx')

sheet = workbook.get_sheet_by_name('Sheet1')

cell = sheet['A1']
cell.value

data = {}
data['fields'] = []


for i in range(1, len(sheet['A'])):
    print(i, sheet.cell(row=i, column=1).value)
    data['fields'].append({
        'name': sheet.cell(row=i, column=1).value
    })

# https://stackoverflow.com/questions/2967194/open-in-python-does-not-create-a-file-if-it-doesnt-exist
with open('simplified.json', 'w+') as json_file:
    json.dump(data, json_file)
